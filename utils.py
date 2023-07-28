import openpyxl
import os
from datetime import date
from openpyxl.styles import Color, Fill
from openpyxl.styles import Font


def get_info_file(file_path):
    #name file
    name_file = file_path.split("-")[-1].split(".")[0]
    print(name_file)
    cpt_update = int(file_path.split("/U")[-1][0])
    flag_first = False
    if cpt_update == 0:
        flag_first = True
    cpt_update += 1
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    #find column
    column_line = 0
    column_linevf = None
    letters =  ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
    for i in letters:
        column_line += 1
        if sheet[i + str(1)].value == "Genosphere Biotechnologies":
            column_linevf = column_line
            break
    if column_linevf is None:
        raise "PB"
    else:
        print("Column to look at:", column_linevf, letters[column_linevf - 1])

    # Update date
    current_date = date.today()
    formatted_dateold = sheet[letters[column_linevf - 1] + "5"].value
    if "/" in formatted_dateold:
        formatted_date = current_date.strftime("%d/%m/%Y")
        LANGUAGE = "FR"
    elif "-" in formatted_dateold:
        formatted_date = current_date.strftime("%d-%m-%Y")
        LANGUAGE = "EN"
    elif "." in formatted_dateold:
        formatted_date = current_date.strftime("%d.%m.%Y")
        LANGUAGE = "GER"
    else:
        raise ValueError("Invalid date format in the Excel file.")
    sheet[letters[column_linevf - 1] + "5"].value = formatted_date

    if flag_first:
        for j in range(2):
            for i in range(5, 100):
                if "Peptide" in str(sheet.cell(row=i, column=column_linevf).value) \
                        or "Peptid" in str(sheet.cell(row=i, column=column_linevf).value):
                        #or '"-"&' in str(sheet.cell(row=i, column=column_linevf).value):
                    sheet.delete_rows(i + 1)

    #for j in range(2):
    for i in range(1, 100):
        #print(str(sheet.cell(row=i, column=column_linevf).value))
        if "*** NEW PURIFICATION REQUIRED ***" in str(sheet.cell(row=i, column=column_linevf).value) \
                or "*** NOUVELLE PURIFICATION REQUISE ***" in str(sheet.cell(row=i, column=column_linevf).value):
            sheet.delete_rows(i)
            #print(str(sheet.cell(row=i, column=column_linevf).value))
            #ok
    #ok

    number_of_line_status = -1
    for i in range(100):
        if sheet[letters[column_linevf-1]+str(i+1)].value == "SUIVI:":
            number_of_line_status = i + 2
        elif sheet[letters[column_linevf - 1] + str(i + 1)].value == "STATUS:":
            number_of_line_status = i + 2
            break

    number_of_product = min(3, (number_of_line_status - 12 - 1))
    if number_of_line_status ==-1 :
        raise "PB"
    else:
        print("suivi/statut line ", number_of_line_status)
        print("nombre de produit", number_of_product)
        print(LANGUAGE)
    #ok
    version_toprint = 0
    step_data2 = [("Add stars to the date", ["NO", "YES"])]
    if LANGUAGE == "FR":
        step_data = [("NOUVELLE TENTATIVE DE SYNTHESE", ["NO", "YES"])]
        possibility_button = ["EN COURS", "EN COURS NPR", "ACHEVEE", "Empty"]
    else:
        step_data = [("NEW SYNTHESIS ATTEMPT", ["NO", "YES"])]
        possibility_button = ["ON GOING", "ON GOING NPR", "COMPLETE", "Empty"]
    #step_data.append(("  ", []))
    for i in range(100):
        #print(sheet[letters[column_linevf - 1] + str(number_of_line_status+1+i)].value)
        if str(sheet[letters[column_linevf - 1] + str(number_of_line_status+1+i)].value) == "DATE ESTIMEE D'ACHEVEMENT:" or \
                str(sheet[letters[column_linevf - 1] + str(number_of_line_status+1+i)].value) == "ANTICIPATED DATE OF COMPLETION :":
            break
        else:
            if sheet[letters[column_linevf - 1] + str(
                    number_of_line_status + 1 + i)].value is not None:
                value = str(sheet[letters[column_linevf-1]+str(number_of_line_status+1+i)].value).split("-")[0] + "-"
                flag_num = False
                print(sheet[letters[column_linevf-1]+str(number_of_line_status+1+i)].value)
                for j in range(1, 100):
                    if str(j) == value[0]:
                        flag_num = True
                        #last = j
                if value != "*** NEW SYNTHESIS ATTEMPT ***-":
                    if flag_num is False:
                        value = str(len(step_data)) + value
                    #print(value)
                if value == "*** NEW SYNTHESIS ATTEMPT ***-":
                    step_data.append((value, []))
                else:
                    step_data.append((value, possibility_button))
    print(step_data)
    #step_data = step_data[:-1]
    #step_data.append(("  ", []))
    #step_data2.append()
    #print(step_data)

    line_date_arrive = -1
    for i in range(1, 100):
        if sheet[letters[column_linevf - 1] + str(i)].value == "DATE ESTIMEE D'ACHEVEMENT:" or \
            sheet[letters[column_linevf - 1] + str( i)].value == "ANTICIPATED DATE OF COMPLETION :":
            line_date_arrive = i
            break
    if line_date_arrive ==-1 :
        raise "PB"
    else:
        print(" line_date_arrive  ", line_date_arrive)

    date_value = sheet[letters[column_linevf - 1] + str(line_date_arrive+1)].value
    print("date value ", date_value)
    if flag_first:
        sheet.delete_rows(line_date_arrive + 2)
        sheet.delete_rows(line_date_arrive - 2)
    if flag_first:
        if LANGUAGE == "FR":
            default_value = ["NO", "EN COURS"] + ["Empty"] * ( len(step_data) -2 ) #+ #+ ["NO"]
        else:
            default_value = ["NO", "ON GOING"] + ["Empty"] * ( len(step_data) -2 ) #+ ["NO"]
    else:
        default_value = []
        flag_goon = False
        if version_toprint > 4:
            num_max = 5
        else:
            num_max = 3
        if LANGUAGE == "FR":
            name_check = "SUIVI:"
        else:
            name_check = "STATUS:"
        if LANGUAGE == "FR":
            for i in range(1, 100):
                #print("OOOOK",sheet[letters[column_linevf - 1] + str(i)].value)
                if flag_goon and len(default_value)<num_max :
                    #print(sheet[letters[column_linevf - 1] + str(i)].value)
                    if str(sheet[letters[column_linevf - 1] + str(i)].value)[-8:] == "EN COURS":
                        default_value.append("EN COURS")
                        #ok
                    elif str(sheet[letters[column_linevf - 1] + str(i)].value)[-7:] == "ACHEVEE":
                        default_value.append("ACHEVEE")
                        #ok
                    else:
                        default_value.append("Empty")
                if sheet[letters[column_linevf - 1] + str(i)].value==name_check:
                    flag_goon = True
                    #print("OK", sheet[letters[column_linevf - 1] + str(i)].value)
        else:
            for i in range(1, 100):
                if flag_goon and len(default_value)<num_max :
                    if str(sheet[letters[column_linevf - 1] + str(i)].value)[-8:] == "ON GOING":
                        default_value.append("ON GOING")
                        #ok
                    elif str(sheet[letters[column_linevf - 1] + str(i)].value)[-7:] == "COMPLETE":
                        default_value.append("COMPLETE")
                        #ok
                    else:
                        default_value.append("Empty")
                if sheet[letters[column_linevf - 1] + str(i)].value==name_check:
                    flag_goon = True

    print("Default value", default_value)
    print()
    return number_of_product, number_of_line_status, column_linevf, sheet, letters, name_file, workbook, version_toprint, \
        cpt_update, step_data, date_value, line_date_arrive, default_value, LANGUAGE, step_data2

def get_files_in_folder(folder_path):
    return [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]

# Function to modify the Excel file

