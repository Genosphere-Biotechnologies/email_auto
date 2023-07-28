from tkinter import *
from datetime import date
from utils import get_info_file, get_files_in_folder
import os
import openpyxl.styles
from tkinter.font import Font  # Import the Font class
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
import time
import glob
class ExcelModifier:
    def __init__(self, file_path, filename, path_to_save):
        self.path_to_save = path_to_save
        self.file_path = file_path
        self.filename = filename
        #print(file_path)
        self.name_file = file_path.split("-")[-1].split(".")[0]

        self.number_of_product, self.number_of_line, self.column_linevf, self.sheet, \
        self.letters, self.name_file, self.workbook, self.version_toprint, self.cpt_update, \
        self.step_data, self.date_value, self.line_date_arrive, self.default_value, self.LANGUAGE \
            = get_info_file(file_path)

        self.name_file_U = "U" + str(self.cpt_update) + "-" + str(name_file) + ".xlsx"
        print(name_file)

        self.label_list_all = []
        self.buttons_list_all = []
        self.label_date_all = []

        self.flag_first_time_line = True


        self.root = Tk()
        self.root.title(" Create Excel " + self.name_file_U) #+ " - Update " + str(self.cpt_update))
        # Set the font size for the labels to 14
        self.label_font = Font(size=30)
        # Set the font size for the buttons to 12
        self.button_font = Font(size=15)


        self.cpt_general = 0
        self.field_all_var = IntVar(value=1)
        self.field_vars = [StringVar() for _ in range(len(self.step_data) * self.number_of_product)]
        for ix, x in enumerate(self.default_value):
            self.field_vars[ix].set(value=x)

    def on_next_button_click(self):
        self.root.destroy()
    def modify_excel(self):
        # Load the Excel file
        offset_line = self.number_of_line + 1
        column_line = self.column_linevf

        # Get user inputs from the selected Radiobuttons for each product
        for index, (label_text, options) in enumerate(self.step_data, start=0):
            selected_value = self.field_vars[index].get()
            if selected_value != "Empty":
                tiret = " -----> "
            else:
                tiret = ""
                selected_value = ""
            #self.sheet.cell(row=index + offset_line, column=column_line + 1, value=selected_value)
            self.sheet.cell(row=index + offset_line, column=column_line, value=label_text + tiret + selected_value)


        self.line_date_arrive = -1
        for i in range(1, 100):
            if self.sheet[self.letters[self.column_linevf - 1] + str(i)].value == "DATE ESTIMEE D'ACHEVEMENT:" or \
                    self.sheet[self.letters[self.column_linevf - 1] + str(i)].value == "ANTICIPATED DATE OF COMPLETION :":
                self.line_date_arrive = i
                break
        if self.line_date_arrive == -1:
            raise "PB"
        else:
            print(" line_date_arrive  ", self.line_date_arrive)
        self.sheet[self.letters[self.column_linevf - 1] + str(self.line_date_arrive+1)].value = self.date_var.get()


        # Save the modified Excel file
        #PATH = f"{self.path_to_save}/{self.name_file}/"
        #if not os.path.exists(PATH):
        #    os.makedirs(PATH)
        modified_file_path = f"{self.path_to_save}/{self.name_file_U}"
        self.workbook.save(modified_file_path)

        print("Excel file modified and saved successfully.")
        os.system(f"open {self.path_to_save}/{self.name_file_U}")
        self.root.destroy()
        #ok



    def modify_excel_multiple(self):
        list_index_1=[]
        if self.LANGUAGE == "FR":
            text_label_pep = "- Peptide "  # + str(peptid+1)
        else:
            text_label_pep = "- Peptide "  # + str(peptid+1)
        #list_index_1.append(text_label_pep)
        for ix, x in  enumerate(self.field_vars):
            peptid = ix // len(self.step_data)
            num_id_here = ix % len(self.step_data)
            if ix % len(self.step_data) == 0:
                print(text_label_pep)
                list_index_1.append(text_label_pep+str(peptid+1))
            if self.LANGUAGE == "FR":
                text_label_pep = "- Peptide " #+ str(peptid+1)
            else:
                text_label_pep = "- Peptid " #+ str(peptid+1)
            offset_line = self.number_of_line + 1
            column_line = self.column_linevf
            selected_value = x.get()
            print(selected_value)
            if selected_value != "Empty":
                tiret = " -----> "
            else:
                tiret = ""
                selected_value = ""
            list_index_1.append(self.step_data[num_id_here][0] + tiret + selected_value)

        offset_line2 = self.number_of_line + 1
        self.sheet.insert_rows(0 + offset_line2)  #+ len(self.step_data))
        for _ in range(((len(self.step_data)+1)*(self.number_of_product-1))):
            self.sheet.insert_rows(0 + offset_line2 + len(self.step_data)+1)

        for ix, x in enumerate(list_index_1):
            self.sheet.cell(row=ix + offset_line, column=column_line, value=x)

        self.line_date_arrive = -1
        for i in range(1, 100):
            if self.sheet[self.letters[self.column_linevf - 1] + str(i)].value == "DATE ESTIMEE D'ACHEVEMENT:" or \
                    self.sheet[
                        self.letters[self.column_linevf - 1] + str(i)].value == "ANTICIPATED DATE OF COMPLETION :":
                self.line_date_arrive = i
                break
        if self.line_date_arrive == -1:
            raise "PB"
        else:
            print(" line_date_arrive  ", self.line_date_arrive)
        self.sheet[
            self.letters[self.column_linevf - 1] + str(self.line_date_arrive + 1)].value = self.date_var.get()

        # Save the modified Excel file
        #PATH = f"{self.path_to_save}/{self.name_file}/"
        #if not os.path.exists(PATH):
        #    os.makedirs(PATH)
        modified_file_path = f"{self.path_to_save}/{self.name_file_U}"
        self.workbook.save(modified_file_path)

        print("Excel file modified and saved successfully.")
        self.root.destroy()



    def update_radiobuttons_visibility(self):
        global labeltest, label0, label1, cpt_general, \
            label2, label_index, label_list, buttons_list, \
            label3, label4, date_label, date_var, date_entry, \
            modify_button, buttons_list_all, label_list_all, label_date_all

        update_all_products = self.field_all_var.get()
        print(update_all_products)
        # for radiobutton in radiobuttons:
        #    radiobutton.grid_forget()

        if update_all_products == 0:
            # Update
            if self.cpt_general > 0:
                self.label0.grid_forget()
                del self.label0
                self.label1.grid_forget()
                del self.label1
                self.label2.grid_forget()
                del self.label2
                self.label3.grid_forget()
                del self.label3
                self.label4.grid_forget()
                del self.label4
                for x in self.label_list:
                    x.grid_forget()
                    del x
                for x in self.buttons_list:
                    for y in x:
                        y.grid_forget()
                        del y
                self.date_label.grid_forget()
                del self.date_label
                # date_var.grid_forget()
                del self.date_var
                self.date_entry.grid_forget()
                del self.date_entry
                self.modify_button.grid_forget()
                del self.modify_button
                self.next_button.grid_forget()
                del self.next_button

            # Plot - difficile
            radiobuttons = []
            self.label_list_all = []
            self.buttons_list_all = []
            self.label_date_all = []
            self.label0 = Label(self.root, text="  ", font=self.label_font)
            self.label0.grid(row=0 + 2, column=0, sticky=W)
            for numpep in range(self.number_of_product):
                offset = 12 * numpep
                self.label1 = Label(self.root, text=" Peptide " + str(numpep + 1), font=self.button_font)
                self.label1.grid(row=0 + 3 + offset, column=0, sticky=W)
                self.label2 = Label(self.root, text="  ", font=self.button_font)
                self.label2.grid(row=0 + 4 + offset, column=0, sticky=W)
                self.label_list = []
                self.buttons_list = []
                for index, (label_text, options) in enumerate(self.step_data):
                    self.label_list.append(Label(self.root, text=label_text, font=self.button_font))
                    self.label_list[-1].grid(row=index + len(self.step_data) + offset, column=0, sticky=W)
                    self.buttons_list.append([])
                    for option_index, option in enumerate(options):
                        self.buttons_list[-1].append(
                            Radiobutton(self.root, text=option, variable=self.field_vars[index + numpep * len(self.step_data)],
                                        value=option, font=self.button_font, height=1, width=10, indicatoron=0))
                        self.buttons_list[-1][-1].grid(row=index + len(self.step_data) + offset, column=option_index + 1, sticky=W)
                        radiobuttons.append(self.buttons_list[-1][-1])
                self.label_list_all.append(self.label_list)
                self.buttons_list_all.append(self.buttons_list)



                self.label4 = Label(self.root, text="  ", font=self.button_font)
                self.label4.grid(row=len(self.step_data) + 7 + offset, column=0, sticky=W)

                #self.label_date_all.append(self.label3)
                self.label_date_all.append(self.label4)
                #self.label_date_all.append(self.date_label)
                #self.label_date_all.append(self.date_entry)
                #self.label_date_all.append(self.label1)

            self.label3 = Label(self.root, text="  ", font=self.button_font)
            self.label3.grid(row=len(self.step_data) + len(self.step_data) + offset, column=0, sticky=W)

            # Add a label and text field for updating the date of receipt
            self.date_label = Label(self.root, text="   Update the date of receipt:", font=self.button_font)
            self.date_label.grid(row=len(self.step_data) + 6 + offset, column=0, sticky=W)

            self.date_var = StringVar()
            self.date_entry = Entry(self.root, textvariable=self.date_var, font=self.button_font)
            self.date_entry.grid(row=len(self.step_data) + 6 + offset, column=1)
            self.date_var.set(self.date_value)

            if update_all_products:
                self.modify_button = Button(self.root, text="Create Excel", command=self.modify_excel, font=self.button_font)
            else:
                self.modify_button = Button(self.root, text="Create Excel", command=self.modify_excel_multiple, font=self.button_font)

            self.modify_button.grid(row=len(self.step_data) + 8 + offset, columnspan=len(self.step_data[0][1]) + 1)

            self.next_button = Button(self.root, text="Next", command=self.on_next_button_click, font=self.button_font)
            self.next_button.grid(row=len(self.step_data) + 9 + offset, columnspan=len(self.step_data[0][1]) + 1)


            self.cpt_general += 1

        else:

            if self.cpt_general > 0:
                self.label0.grid_forget()
                del self.label0
                self.label1.grid_forget()
                del self.label1
                self.label2.grid_forget()
                del self.label2
                self.label3.grid_forget()
                del self.label3
                self.label4.grid_forget()
                del self.label4
                for y in self.label_list_all:
                    for x in y:
                        x.grid_forget()
                        del x
                for z in self.buttons_list_all:
                    for x in z:
                        for y in x:
                            y.grid_forget()
                            del y
                for x in self.label_date_all:
                    x.grid_forget()
                    del x

                self.date_label.grid_forget()
                del self.date_label
                # date_var.grid_forget()
                del self.date_var
                self.date_entry.grid_forget()
                del self.date_entry
                self.modify_button.grid_forget()
                del self.modify_button
                self.next_button.grid_forget()
                del self.next_button

            radiobuttons = []
            self.label0 = Label(self.root, text="  ", font=self.button_font)
            self.label0.grid(row=0 + 2, column=0, sticky=W)
            self.label1 = Label(self.root, text=" All Peptides ", font=self.button_font)
            self.label1.grid(row=0 + 3, column=0, sticky=W)
            self.label2 = Label(self.root, text="  ", font=self.button_font)
            self.label2.grid(row=0 + 4, column=0, sticky=W)
            self.label_list = []
            self.buttons_list = []
            for index, (label_text, options) in enumerate(self.step_data):
                self.label_list.append(Label(self.root, text=label_text, font=self.button_font))
                self.label_list[-1].grid(row=index + len(self.step_data), column=0, sticky=W)
                self.buttons_list.append([])
                for option_index, option in enumerate(options):
                    self.buttons_list[-1].append(Radiobutton(self.root, text=option, variable=self.field_vars[index],
                                                             value=option, font=self.button_font, height=2, width=10, indicatoron=0))
                    self.buttons_list[-1][-1].grid(row=index + len(self.step_data), column=option_index + 1, sticky=W)
                    radiobuttons.append(self.buttons_list[-1][-1])

            self.label3 = Label(self.root, text="  ")
            self.label3.grid(row=len(self.step_data) + len(self.step_data), column=0, sticky=W)

            # Add a label and text field for updating the date of receipt
            self.date_label = Label(self.root, text=" Update the date of receipt: ", font=self.button_font)
            self.date_label.grid(row=len(self.step_data) + 6, column=0, sticky=W)

            self.date_var = StringVar()
            self.date_entry = Entry(self.root, textvariable=self.date_var, font=self.button_font)
            self.date_entry.grid(row=len(self.step_data) + 6, column=1)
            self.date_var.set(self.date_value)

            self.label4 = Label(self.root, text="  ")
            self.label4.grid(row=len(self.step_data) + 7, column=0, sticky=W)

            # Create a button to trigger the modification process
            if self.field_all_var.get():
                self.modify_button = Button(self.root, text="Create Excel", command=self.modify_excel,
                                            font=self.button_font)
            else:
                self.modify_button = Button(self.root, text="Create Excel", command=self.modify_excel_multiple,
                                            font=self.button_font)

            self.modify_button.grid(row=len(self.step_data) + 8 , columnspan=len(self.step_data[0][1]) + 1)

            self.next_button = Button(self.root, text="Next", command=self.on_next_button_click, font=self.button_font)
            self.next_button.grid(row=len(self.step_data) + 9, columnspan=len(self.step_data[0][1]) + 1)

            self.cpt_general += 1
    def start_main(self):
        # Implementation of the start_main function as in the original code
        global cpt_general, field_all_var, label0, label1, label2, label_list, buttons_list, label3, date_label, label4, date_var, date_entry, \
            modify_button, field_vars

        cpt_general = 0
        # Create a checkbox to ask if the update is for all products
        self.field_all_var = IntVar(value=1)
        field_all_checkbox = Checkbutton(self.root, text="Update all products with the same status",
                                         variable=self.field_all_var, font=self.button_font, indicatoron=0)
        # Set the font size for the buttons to 12
        field_all_checkbox.grid(row=0, columnspan=2)

        # Create StringVars to store the selected values for Radiobuttons
        self.field_vars = [StringVar() for _ in range(len(self.step_data * self.number_of_product))]
        for i in range(self.number_of_product):
            for ix, x in enumerate(self.default_value):
                self.field_vars[ix+i*len(self.step_data)].set(value=x)
        radiobuttons = []
        self.label0 = Label(self.root, text="  ", font=self.button_font)
        self.label0.grid(row=0 + 2, column=0, sticky=W)
        self.label1 = Label(self.root, text=" All Peptides ", font=self.button_font)
        self.label1.grid(row=0 + 3, column=0, sticky=W)
        self.label2 = Label(self.root, text="  ", font=self.button_font)
        self.label2.grid(row=0 + 4, column=0, sticky=W)
        self.label_list = []
        self.buttons_list = []
        for index, (label_text, options) in enumerate(self.step_data):
            self.label_list.append(Label(self.root, text=label_text, font=self.button_font))
            self.label_list[-1].grid(row=index + len(self.step_data), column=0, sticky=W)
            self.buttons_list.append([])
            for option_index, option in enumerate(options):
                self.buttons_list[-1].append(Radiobutton(self.root, text=option, variable=self.field_vars[index],
                                                         value=option, font=self.button_font,
                                                         height=2, width=10, indicatoron=0))
                self.buttons_list[-1][-1].grid(row=index + len(self.step_data), column=option_index + 1, sticky=W)
                radiobuttons.append(self.buttons_list[-1][-1])

        self.label3 = Label(self.root, text="  ", font=self.button_font)
        self.label3.grid(row=len(self.step_data) + len(self.step_data), column=0, sticky=W)

        # Add a label and text field for updating the date of receipt
        self.date_label = Label(self.root, text=" Update the date of receipt: ", font=self.button_font)
        self.date_label.grid(row=len(self.step_data) + 6, column=0, sticky=W)

        self.date_var = StringVar()
        self.date_entry = Entry(self.root, textvariable=self.date_var, font=self.button_font)
        self.date_var.set(self.date_value)
        self.date_entry.grid(row=len(self.step_data) + 6, column=1)

        self.label4 = Label(self.root, text="  ", font=self.button_font)
        self.label4.grid(row=len(self.step_data) + 7, column=0, sticky=W)

        # Create a button to trigger the modification process
        print(self.field_all_var.get())
        if self.field_all_var.get()==1:
            self.modify_button = Button(self.root, text="Create Excel", command=self.modify_excel,
                                        font=self.button_font)
        else:
            self.modify_button = Button(self.root, text="Create Excel", command=self.modify_excel_multiple,
                                        font=self.button_font)

        self.modify_button.grid(row=len(self.step_data) + 8, columnspan=len(self.step_data[0][1]) + 1)

        self.next_button = Button(self.root, text="Next", command=self.on_next_button_click, font=self.button_font)
        self.next_button.grid(row=len(self.step_data) + 9, columnspan=len(self.step_data[0][1]) + 1)

        self.cpt_general += 1

        # Bind the checkbox to the function to update the Radiobuttons visibility
        field_all_checkbox.config(command=self.update_radiobuttons_visibility)

        self.root.mainloop()

    def run(self):
        self.start_main()
        #self.root.mainloop()




if __name__ == "__main__":

    folder_path = "../U0_Updates/"  # Change this to your desired folder path
    files = get_files_in_folder(folder_path)
    timestr = time.strftime("%Y_%m_%d")#_%H_%M_%S")

    PATH = f"../Un_Updates/{timestr}/"
    if not os.path.exists(PATH):
        os.makedirs(PATH)


    txtfiles = []
    for file in glob.glob("../Un_Updates/*/*.xlsx"):
        print(file)
        txtfiles.append(file)
    print(txtfiles)
    #ok

    for filename in files:
        file_path = folder_path + filename
        name_file = file_path.split("-")[-1].split(".")[0]
        #print(name_file)
        #test Un
        update = None
        for i in range(100, 0, -1):
            name_file_U = "U"+str(i)+"-"+str(name_file)+".xlsx"
            #print(name_file_U, txtfiles)
            for x in txtfiles:
                if name_file in x:
                    update = x
        #print(update)
        if update is None:
            #print("U"+str(0)+"-"+str(name_file)+".xlsx")
            excel_modifier = ExcelModifier(file_path,  "U"+str(0)+"-"+str(name_file)+".xlsx", PATH)
            excel_modifier.run()
        else:
            #print(update, name_file_U)
            #ok
            excel_modifier = ExcelModifier(update, name_file_U, PATH)
            excel_modifier.run()
